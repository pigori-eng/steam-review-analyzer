"""
Steam Review Analyzer API v3.0
FastAPI 백엔드 - 전체 엔드포인트
"""
import json
import os
import io
import csv
import asyncio
from typing import Optional
from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel

from crawler import SteamCrawler
from analyzer import ReviewAnalyzer
from sample_data import SAMPLE_REVIEWS, SAMPLE_GAME_INFO

app = FastAPI(title="Steam Review Analyzer API", version="3.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── 전역 태스크 저장소 ──
tasks: dict = {}


# ── Request Models ──
class AnalyzeRequest(BaseModel):
    url: str
    review_count: int = 100
    day_range: Optional[int] = None
    api_key: Optional[str] = None
    ai_provider: str = "openai"

class ExportRequest(BaseModel):
    task_id: str
    format: str = "csv"  # csv | xlsx | txt


# ─────────────────────────────────────
# Health & Root
# ─────────────────────────────────────
@app.get("/")
def root():
    return {"name": "Steam Review Analyzer API", "version": "3.0.0", "status": "ok"}

@app.get("/api/health")
def health():
    return {"status": "healthy", "version": "3.0.0"}


# ─────────────────────────────────────
# Sample Data
# ─────────────────────────────────────
@app.get("/api/sample")
def get_sample():
    analyzer = ReviewAnalyzer()
    result = analyzer.analyze(SAMPLE_REVIEWS, SAMPLE_GAME_INFO)
    return {
        "game_info": SAMPLE_GAME_INFO,
        "statistics": result["statistics"],
        "category_analysis": result["category_analysis"],
        "language_analysis": result["language_analysis"],
        "playtime_analysis": result["playtime_analysis"],
        "time_trends": result["time_trends"],
        "analysis": result["ai_analysis"],
        "reviews": SAMPLE_REVIEWS[:100],
        "note": result.get("note")
    }


# ─────────────────────────────────────
# Main Analyze (Sync)
# ─────────────────────────────────────
@app.post("/api/analyze")
async def analyze(req: AnalyzeRequest):
    try:
        # 1. Crawl
        crawler = SteamCrawler()
        crawl_result = crawler.crawl(req.url, count=req.review_count, day_range=req.day_range)
        if not crawl_result:
            raise HTTPException(status_code=400, detail="크롤링 실패")

        game_info = crawl_result.get("game_info", {})
        reviews = crawl_result.get("reviews", [])

        if not reviews:
            raise HTTPException(status_code=404, detail="리뷰를 찾을 수 없습니다")

        # 2. Analyze
        analyzer = ReviewAnalyzer(api_key=req.api_key, ai_provider=req.ai_provider)
        result = analyzer.analyze(reviews, game_info)

        return {
            "game_info": game_info,
            "statistics": result["statistics"],
            "category_analysis": result["category_analysis"],
            "language_analysis": result["language_analysis"],
            "playtime_analysis": result["playtime_analysis"],
            "time_trends": result["time_trends"],
            "analysis": result["ai_analysis"],
            "reviews": reviews[:200],
            "note": result.get("note")
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────
# Async Analyze (Background Task)
# ─────────────────────────────────────
@app.post("/api/analyze/async")
async def analyze_async(req: AnalyzeRequest, background_tasks: BackgroundTasks):
    import uuid
    task_id = str(uuid.uuid4())
    tasks[task_id] = {"status": "pending", "progress": 0, "message": "분석 대기 중...", "result": None}
    background_tasks.add_task(_run_analysis, task_id, req)
    return {"task_id": task_id, "status": "started"}


async def _run_analysis(task_id: str, req: AnalyzeRequest):
    try:
        tasks[task_id]["status"] = "running"
        tasks[task_id]["progress"] = 10
        tasks[task_id]["message"] = "크롤링 시작..."

        crawler = SteamCrawler()
        crawl_result = crawler.crawl(req.url, count=req.review_count, day_range=req.day_range)

        if not crawl_result:
            tasks[task_id]["status"] = "error"
            tasks[task_id]["message"] = "크롤링 실패"
            return

        tasks[task_id]["progress"] = 40
        tasks[task_id]["message"] = "AI 분석 시작..."

        game_info = crawl_result.get("game_info", {})
        reviews = crawl_result.get("reviews", [])

        analyzer = ReviewAnalyzer(api_key=req.api_key, ai_provider=req.ai_provider)
        result = analyzer.analyze(reviews, game_info)

        tasks[task_id]["progress"] = 100
        tasks[task_id]["status"] = "completed"
        tasks[task_id]["message"] = "분석 완료"
        tasks[task_id]["result"] = {
            "game_info": game_info,
            "statistics": result["statistics"],
            "category_analysis": result["category_analysis"],
            "language_analysis": result["language_analysis"],
            "playtime_analysis": result["playtime_analysis"],
            "time_trends": result["time_trends"],
            "analysis": result["ai_analysis"],
            "reviews": reviews[:200],
            "note": result.get("note")
        }
    except Exception as e:
        tasks[task_id]["status"] = "error"
        tasks[task_id]["message"] = str(e)


@app.get("/api/task/{task_id}")
def get_task(task_id: str):
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="태스크를 찾을 수 없습니다")
    return tasks[task_id]


# ─────────────────────────────────────
# Export API
# ─────────────────────────────────────
@app.post("/api/export")
async def export_data(req: ExportRequest):
    task = tasks.get(req.task_id)
    if not task or task.get("status") != "completed":
        raise HTTPException(status_code=400, detail="완료된 분석 결과가 없습니다")

    result = task["result"]
    reviews = result.get("reviews", [])
    game_name = result.get("game_info", {}).get("name", "game")

    if req.format == "csv":
        return _export_csv(reviews, game_name)
    elif req.format == "xlsx":
        return _export_xlsx(reviews, result, game_name)
    elif req.format == "txt":
        return _export_txt(result, game_name)
    else:
        raise HTTPException(status_code=400, detail="지원하지 않는 형식입니다")


def _export_csv(reviews: list, game_name: str) -> StreamingResponse:
    output = io.StringIO()
    if not reviews:
        output.write("No data\n")
    else:
        fields = ["author_id", "language", "playtime_hours", "voted_up", "votes_up", "review", "timestamp_created"]
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for r in reviews:
            writer.writerow({k: r.get(k, "") for k in fields})

    output.seek(0)
    filename = f"{game_name}_reviews.csv".replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _export_xlsx(reviews: list, result: dict, game_name: str) -> StreamingResponse:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()

        # Sheet 1: Reviews
        ws1 = wb.active
        ws1.title = "Reviews"
        headers = ["ID", "Language", "Playtime(h)", "Positive", "Votes", "Review", "Date"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F6FEB")
            cell.alignment = Alignment(horizontal="center")

        for row, r in enumerate(reviews, 2):
            import datetime
            ts = r.get("timestamp_created", 0)
            date_str = datetime.datetime.fromtimestamp(ts).strftime("%Y-%m-%d") if ts else ""
            ws1.cell(row=row, column=1, value=r.get("author_id", ""))
            ws1.cell(row=row, column=2, value=r.get("language", ""))
            ws1.cell(row=row, column=3, value=round(r.get("playtime_hours", 0), 1))
            ws1.cell(row=row, column=4, value="긍정" if r.get("voted_up") else "부정")
            ws1.cell(row=row, column=5, value=r.get("votes_up", 0))
            ws1.cell(row=row, column=6, value=r.get("review", "")[:500])
            ws1.cell(row=row, column=7, value=date_str)

        # Sheet 2: Statistics
        ws2 = wb.create_sheet("Statistics")
        stats = result.get("statistics", {})
        ws2.cell(row=1, column=1, value="항목").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="수치").font = Font(bold=True)
        for i, (k, v) in enumerate(stats.items(), 2):
            ws2.cell(row=i, column=1, value=k)
            ws2.cell(row=i, column=2, value=str(v) if isinstance(v, dict) else v)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{game_name}_analysis.xlsx".replace(" ", "_")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl이 설치되지 않았습니다")


def _export_txt(result: dict, game_name: str) -> StreamingResponse:
    lines = [
        f"=== {game_name} Steam 리뷰 분석 보고서 ===\n",
        f"총 리뷰: {result.get('statistics', {}).get('total_reviews', 0)}건",
        f"긍정률: {result.get('statistics', {}).get('positive_rate', 0)}%",
        f"평균 플레이타임: {result.get('statistics', {}).get('avg_playtime', 0)}h\n",
    ]
    ai = result.get("analysis", {})
    exec_d = ai.get("executive_dashboard", {})
    for insight in exec_d.get("ai_insights", []):
        lines.append(f"[{insight.get('priority','')}] {insight.get('title','')}")
        lines.append(f"  근거: {insight.get('evidence','')}")
        lines.append(f"  액션: {insight.get('action_plan','')}\n")

    content = "\n".join(lines)
    filename = f"{game_name}_report.txt".replace(" ", "_")
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=False)
